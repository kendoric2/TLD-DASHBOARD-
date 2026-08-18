"""
Loop every egress endpoint, pull ALL column names + a sample value for each,
and export to Excel — one sheet per endpoint (sheet named after the endpoint),
with two columns: "Column Name" and "Sample Value".

Built to see, at a glance, exactly what data each endpoint exposes — e.g. for
finding the right columns for vendor-performance tracking.

Sample values are pulled in small COLUMN BATCHES (wide endpoints like policies
have ~2000 columns and the API rejects asking for them all at once).

Requires openpyxl:   pip install openpyxl

Run everything (slow — policies alone is ~2000 columns):
    python3 sandbox/probes/probe_all_columns.py

Run only certain endpoints and MERGE them into the existing file (fast — use this
when new endpoints get enabled):
    python3 sandbox/probes/probe_all_columns.py dialer_leads lead_logs vendor_logs

Output:               egress_columns.xlsx  (saved in the project root)
"""
import os
import sys
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "src"))
import json
import _probe_lib as p

config = p.config

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    raise SystemExit("openpyxl isn't installed.  Run:  pip install openpyxl")

config.require_creds()

# ---- settings you can tweak ----
ALL_ENDPOINTS = ["policies", "leads", "agentcpa", "vendors", "users",
                 "vendorperformance", "report_cpa_agent", "tldialer/report_agentcpa",
                 # enabled later — the dialer/log side
                 "dialer_leads", "lead_dialer_leads", "lead_logs", "vendor_logs"]
# endpoints named on the command line -> do only those and merge into the existing file
ENDPOINTS = sys.argv[1:] or ALL_ENDPOINTS
MERGE = bool(sys.argv[1:])
BATCH = 100          # how many columns to request per call (wide endpoints)
SAMPLE_ROWS = 3      # rows pulled per batch to find a non-null value
# sensitive substrings — show a placeholder instead of the real value
MASK = ("ssn", "cc_number", "cc_cvv", "cc_exp", "bank_account", "bank_routing",
        "mothers_maiden", "drivers_license", "passport", "medicare_claim",
        "medicaid_username", "medicaid_password", "dob", "cms_password",
        "healthsherpa_password", "vicidial_pass", "personal_")


def sheet_name(ep):
    name = ep
    for ch in "[]:*?/\\":
        name = name.replace(ch, "_")
    return name[:31]


def first_value(col, rows):
    cl = col.lower()
    for r in rows:
        v = r.get(col)
        if v in p.EMPTY:
            continue
        if any(m in cl for m in MASK):
            return "*** redacted ***"
        if isinstance(v, (dict, list)):
            v = json.dumps(v)
        return str(v)[:500]
    return ""


def columns_and_values(ep):
    """Return (column_names, {col: sample_value}). Handles report endpoints whose
    /docs/columns returns data rows, and batches wide endpoints.

    NOTE: `<endpoint>/docs/columns` is permissioned SEPARATELY from the endpoint itself,
    so it can be blocked even when the endpoint works. When that happens we fall back to
    sampling a row and using its keys — that's only TLD's default subset, not every
    column, but it beats an empty sheet."""
    try:
        docs = config.egress_get(f"{ep}/docs/columns")
    except Exception as ex:
        print(f"      /docs/columns unavailable ({type(ex).__name__})")
        docs = None

    if isinstance(docs, list) and docs and isinstance(docs[0], dict):
        cols = list(docs[0].keys())                       # report: docs IS data
        return cols, {c: first_value(c, docs) for c in cols}

    cols = [c for c in docs if isinstance(c, str)] if isinstance(docs, list) else []

    if not cols:
        note = docs if isinstance(docs, dict) else "no column list returned"
        print(f"      /docs/columns not usable ({str(note)[:80]}) — sampling a row instead")
        try:
            rows, _ = p.as_rows(config.egress_get(ep, {"limit": 3}, timeout=60))
        except Exception as ex:
            print(f"      sampling failed too ({type(ex).__name__}) — skipping")
            return [], {}
        if rows and isinstance(rows[0], dict):
            cols = sorted(rows[0].keys())
            print(f"      recovered {len(cols)} column(s) from a sample row "
                  f"(enable {ep}/docs/columns for the full list)")
            return cols, {c: first_value(c, rows) for c in cols}
        return [], {}
    values = {}
    for i in range(0, len(cols), BATCH):
        chunk = cols[i:i + BATCH]
        try:
            rows, _ = p.as_rows(config.egress_get(ep, {"columns": chunk, "limit": SAMPLE_ROWS}, timeout=60))
        except Exception:
            rows = []
        for c in chunk:
            values[c] = first_value(c, rows)
        got = sum(1 for c in chunk if values[c])
        print(f"      cols {i + 1}-{min(i + BATCH, len(cols))} of {len(cols)}  ({got} with values)", flush=True)
    return cols, values


OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
                   "egress_columns.xlsx")

# When specific endpoints are named, keep every sheet already in the file and just
# refresh the ones we're re-pulling — so a targeted run can't wipe the reference.
if MERGE and os.path.exists(OUT):
    from openpyxl import load_workbook
    wb = load_workbook(OUT)
    print(f"merging into existing {os.path.basename(OUT)} "
          f"({len(wb.sheetnames)} sheets already there)")
else:
    wb = Workbook()
    wb.remove(wb.active)

for ep in ENDPOINTS:
    print(f"... {ep}", flush=True)
    try:
        cols, values = columns_and_values(ep)
    except Exception as ex:
        print(f"    skipped ({type(ex).__name__})")
        cols, values = [], {}

    sn = sheet_name(ep)
    if sn in wb.sheetnames:                 # refreshing an endpoint we already had
        wb.remove(wb[sn])
    ws = wb.create_sheet(sn)
    ws.append(["Column Name", "Sample Value"])
    ws["A1"].font = ws["B1"].font = Font(bold=True)
    for c in cols:
        ws.append([c, values.get(c, "")])
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 52
    ws.freeze_panes = "A2"
    print(f"    {len(cols)} columns, {sum(1 for c in cols if values.get(c))} with sample values")

wb.save(OUT)
print(f"\nSaved -> {OUT}")
print(f"Sheets: {', '.join(wb.sheetnames)}")
