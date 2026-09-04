"""
Plain Excel export for the Agent Detail view.

Just the data: a header row and the rows underneath, same eight columns as the screen and
in the same order you've sorted them. No styling, no summary block. Built in memory and
streamed straight to the browser.
"""
import io

from openpyxl import Workbook

# Exactly the columns shown in the Agent Detail table: (header, row key)
COLUMNS = [
    ("Date Sold", "date_sold"),
    ("Lead ID",   "lead_id"),
    ("Role",      "role"),
    ("Agent",     "agent"),
    ("Enroller",  "enroller"),
    ("State",     "state"),
    ("Carrier",   "carrier"),
    ("Plan",      "plan"),
    ("SEP",       "sep"),
]


def sort_rows(rows, key="date_sold", desc=True):
    """Same ordering as the on-screen table: blanks sink to the bottom either direction."""
    keys = {c[1] for c in COLUMNS}
    if key not in keys:
        key = "date_sold"
    filled = [r for r in rows if str(r.get(key) or "").strip()]
    blank = [r for r in rows if not str(r.get(key) or "").strip()]
    if key == "lead_id":
        filled.sort(key=lambda r: float(str(r.get(key) or 0) or 0), reverse=desc)
    else:
        filled.sort(key=lambda r: str(r.get(key) or "").lower(), reverse=desc)
    return filled + blank


def safe_name(s):
    """'Powers, Tony' -> 'Powers-Tony' so filenames stay clean."""
    out = "".join(ch if ch.isalnum() else "-" for ch in str(s or "").strip())
    while "--" in out:
        out = out.replace("--", "-")
    return out.strip("-") or "agent"


def _sheet_title(agent):
    """Excel tab names: max 31 chars, and []:*?/\\ aren't allowed."""
    t = "".join(" " if ch in "[]:*?/\\" else ch for ch in str(agent or "")).strip()
    return (t[:31] or "Agent Detail")


BILLED_COLUMNS = [
    ("Call Time",   "call_date"),
    ("Vendor",      "vendor"),
    ("Agent",       "agent"),
    ("Disposition", "status"),
    ("Talk (sec)",  "talk_sec"),
    ("Cost",        "cost"),
    ("Lead ID",     "lead_id"),        # the CRM lead id (from lead_vendor_lead_code)
    ("Dialer ID",   "dialer_lead_id"), # VICIdial's own id, for TLD support conversations
]


def build_dispo(title, notes, rows, total, tab_name="Call Dispositions"):
    """Call dispositions as a plain sheet: what it covers, then Disposition / Count / Share.

    `notes` are the context lines that make the numbers interpretable later — the range, the
    direction, and how many robocalls were excluded. Without those, a share column is
    meaningless six months from now."""
    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title(tab_name)

    ws.append([title])
    for n in notes or []:
        ws.append([n])
    ws.append([])
    ws.append(["Disposition", "Count", "Share"])

    for r in rows:
        count = r.get("count") or 0
        ws.append([r.get("status") or "",
                   count,
                   round(count / total * 100, 1) if total else 0])
    if rows:
        ws.append([])
        ws.append(["TOTAL", total, 100.0 if total else 0])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_billed(start, end, vendor_label, summary, rows):
    """Invoice audit: the calls billed in a range, one per row, for checking against a
    vendor's bill. Header names the vendor and range, then the totals, then the calls."""
    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title(vendor_label or "Billed Calls")

    s = summary or {}
    ws.append([f"Billed calls: {vendor_label or 'all vendors'}"])
    ws.append([f"Range: {start} to {end}"])
    ws.append([f"{s.get('calls', 0)} calls  ·  ${s.get('spend', 0):,.2f} total  ·  "
               f"{s.get('sales', 0)} sales  ·  "
               f"{s.get('dropped', 0)} unanswered (${s.get('dropped_cost', 0):,.2f})"])
    ws.append([])
    ws.append([h for h, _k in BILLED_COLUMNS])

    for row in rows:
        line = []
        for _h, key in BILLED_COLUMNS:
            val = row.get(key)
            val = "" if val is None else val
            if key in ("cost", "talk_sec") and str(val).strip():
                try:
                    val = float(val)
                except (TypeError, ValueError):
                    pass
            if key in ("lead_id", "dialer_lead_id") and str(val).strip():
                try:
                    val = int(str(val))
                except (TypeError, ValueError):
                    pass
            line.append(val)
        ws.append(line)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, f"BilledCalls_{safe_name(vendor_label or 'all')}_{start}_{end}.xlsx"


def build(agent, start, end, range_label, summary, rows):
    """Return (BytesIO, filename): who it's for, then the header row and their deals."""
    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title(agent)

    # who this report is for (the Agent column shows the closer, which isn't the same
    # person on rows they only enrolled) + the range, then a blank line before the table
    ws.append([f"Agent: {agent}"])
    ws.append([f"Range: {start} to {end}"])
    ws.append([])
    ws.append([h for h, _k in COLUMNS])
    for row in rows:
        line = []
        for _h, key in COLUMNS:
            val = row.get(key)
            val = "" if val is None else val
            if key == "lead_id" and str(val).strip():
                try:
                    val = int(str(val))          # keep lead ids numeric so Excel sorts them right
                except (TypeError, ValueError):
                    pass
            line.append(val)
        ws.append(line)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, f"AgentDetail_{safe_name(agent)}_{start}_{end}.xlsx"


ACTIVE_COLUMNS = [
    ("Carrier",   "carrier"),
    ("Active",    "active"),
    ("Sold",      "sold"),
    ("% Active",  "pct_active"),
]


def build_active(start, end, range_label, rows):
    """Active Policies, one row per carrier — same numbers as the on-screen table
    (active vs. sold, of what was sold in the selected range)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Active Policies"

    ws.append([f"Active Policies — {range_label}"])
    ws.append([f"Range: {start} to {end}"])
    ws.append(["In force, of what was sold in this range"])
    ws.append([])
    ws.append([h for h, _k in ACTIVE_COLUMNS])

    total_active = total_sold = 0
    for r in rows:
        active, sold = r.get("active") or 0, r.get("sold") or 0
        pct = round(active / sold * 100, 1) if sold else 0
        ws.append([r.get("carrier") or "", active, sold, pct])
        total_active += active
        total_sold += sold
    if rows:
        ws.append([])
        pct = round(total_active / total_sold * 100, 1) if total_sold else 0
        ws.append(["TOTAL", total_active, total_sold, pct])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, f"ActivePolicies_{start}_{end}.xlsx"


def build_state(start, end, range_label, rows, include_agents=False):
    """Production by State: one row per state x carrier (GTL-excluded, same as the on-screen
    map). Agent breakdown is a second sheet, added only when the caller opts in — most people
    only care about the carrier split."""
    wb = Workbook()
    ws = wb.active
    ws.title = "By Carrier"

    ws.append([f"Production by State — {range_label}"])
    ws.append([f"Range: {start} to {end}"])
    ws.append(["Deals by customer state, broken down by carrier (GTL excluded)"])
    ws.append([])
    ws.append(["State", "Carrier", "Deals", "State Total", "Share of State"])
    for s in rows:
        state_total = s.get("count") or 0
        for c in (s.get("carriers") or []):
            cnt = c.get("count") or 0
            share = round(cnt / state_total * 100, 1) if state_total else 0
            ws.append([s.get("state") or "", c.get("label") or "", cnt, state_total, share])

    if include_agents:
        ws2 = wb.create_sheet("By Agent")
        ws2.append([f"Production by State — {range_label} (by agent)"])
        ws2.append([f"Range: {start} to {end}"])
        ws2.append([])
        ws2.append(["State", "Agent", "Deals", "State Total", "Share of State"])
        for s in rows:
            state_total = s.get("count") or 0
            for a in (s.get("agents") or []):
                cnt = a.get("count") or 0
                share = round(cnt / state_total * 100, 1) if state_total else 0
                ws2.append([s.get("state") or "", a.get("name") or "", cnt, state_total, share])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    suffix = "_with_agents" if include_agents else ""
    return buf, f"ProductionByState_{start}_{end}{suffix}.xlsx"
